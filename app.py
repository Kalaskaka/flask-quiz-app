from flask import Flask, request, render_template_string, redirect, url_for, session, send_file
import os
import re
import openpyxl
from openpyxl import Workbook
from datetime import datetime
from io import BytesIO

ADMIN_PASSWORD = "1034"  # 任意の安全なパスワードに変更してください

app = Flask(__name__)
app.secret_key = "f3a97cc8b2145"  # 安全なランダム文字列に変更推奨
RESULTS_FILE = "results.xlsx"

correct_answers = {
    "q1": "4", "q2": "2", "q3": "5", "q4": "2", "q5": "1",
    "q6": "1", "q7": "1", "q8": "2", "q9": "5"
}

question_texts = {
    "q1": "問1. 乳牛の妊娠期間（日）と平均乳量（kg/年）",
    "q2": "問2. 牛は幾つの胃を持っているのか？",
    "q3": "問3. 飼料中の繊維・デンプンは何に変わるか？",
    "q4": "問4. ルーメンで合成されるビタミン類の組み合わせ",
    "q5": "問5. 濃厚飼料多給時に、粗飼料を先に給与することでルーメンpHの過度な低下を防ぐことができるか？",
    "q6": "問6. 濃厚飼料を先に給与すると、ルーメンpHは過度に低下する危険があるか？",
    "q7": "問7. TMR（完全混合飼料）を給与することで、ルーメンpHの安定に役立つか？",
    "q8": "問8. 乾草の水分含量は何％未満か？",
    "q9": "問9. 搾乳する施設の名称は？",
    "q10": "問10. あなたが思い描く将来の国内畜産物（肉・乳）像はどのようなものか記述しなさい。（400字以内）"
}

choice_texts = {
    "q1": {
        "1": "妊娠期間 80日 ・ 平均乳量6,000kg/年",
        "2": "妊娠期間180日 ・ 平均乳量4,000kg/年",
        "3": "妊娠期間280日 ・ 平均乳量900kg/年",
        "4": "妊娠期間280日 ・ 平均乳量9,000kg/年",
        "5": "妊娠期間280日 ・ 平均乳量90,000kg/年"
    },
    "q2": {"1": "5つ", "2": "4つ", "3": "3つ", "4": "2つ", "5": "1つ"},
    "q3": {"1": "粗灰分", "2": "ビタミン", "3": "タンパク質", "4": "エタノール", "5": "揮発性脂肪酸"},
    "q4": {"1": "ビタミンA・B・ビオチン", "2": "ビタミンA・B・K", "3": "ビタミンB・C・K", "4": "ビタミンB・D・K", "5": "ビタミンA・B・葉酸"},
    "q5": {"1": "正しい", "2": "誤り"},
    "q6": {"1": "正しい", "2": "誤り"},
    "q7": {"1": "正しい", "2": "誤り"},
    "q8": {"1": "5%", "2": "15%", "3": "30%", "4": "60%", "5": "90%"},
    "q9": {"1": "バルククーラー", "2": "フリーバーン", "3": "繁殖牛舎", "4": "トレンチサイロ", "5": "ミルキングパーラー"}
}

def admin_login_form():
    return '''
        <h2>管理者ログイン</h2>
        <form method="post">
            管理者パスワード：<input type="password" name="password" required>
            <input type="submit" value="ログイン">
        </form>
    '''

index_template = """
<!DOCTYPE html>
<html>
<head>
  <meta charset="utf-8">
  <title>小テスト開始</title>
  <style>
    body {
      line-height: 1.5;
      font-size: 16px;
    }
    .form-field {
      margin-bottom: 12px;
    }
    .submit-button {
      font-size: 18px;
      padding: 10px 20px;
      margin-top: 20px;
    }
  </style>
</head>
<body>
  <h2>学生情報を入力してください</h2>
  <p style="color: red; font-weight: bold;">
    回答は1回限りです。注意深く回答してください。
  </p>
  <form action="/start" method="POST">
    <div class="form-field">
      学籍番号：<input type="text" name="student_id" required>
    </div>
    <div class="form-field">
      氏名：<input type="text" name="name" required>
    </div>
    <div class="form-field">
      メールアドレス：<input type="email" name="email" required>
    </div>
    <input type="submit" value="回答へ進む" class="submit-button">
  </form>
</body>
</html>
"""

question_template = """
<!DOCTYPE html>
<html>
<head><meta charset="utf-8"><title>設問</title></head>
<body>
  <h2>設問に答えてください</h2>
  <form action="/submit" method="POST">
    {% for qkey, qtext in question_texts.items() %}
      <p><strong>{{ qtext }}</strong></p>
      {% if qkey != 'q10' %}
        {% for val, text in choice_texts[qkey].items() %}
          <label><input type="radio" name="{{ qkey }}" value="{{ val }}" required> {{ text }}</label><br>
        {% endfor %}
      {% else %}
        <textarea name="q10" rows="6" cols="60" maxlength="400" required></textarea>
      {% endif %}
    {% endfor %}
    <br><input type="submit" value="送信">
  </form>
</body>
</html>
"""

@app.route("/")
def index():
    return render_template_string(index_template)

@app.route("/start", methods=["POST"])
def start():
    student_id = request.form["student_id"]
    if not (student_id.startswith("542") and len(student_id) == 8):
        return "この学籍番号は無効です。あなたの学籍番号を入力して下さい"
    session["student_id"] = student_id
    session["name"] = request.form["name"]
    session["email"] = request.form["email"]
    if os.path.exists(RESULTS_FILE):
        wb = openpyxl.load_workbook(RESULTS_FILE)
        ws = wb.active
        if any(row[0].value == student_id and student_id != "54200000" for row in ws.iter_rows(min_row=2)):
            return "この学籍番号では既に受験済みです。"
    return render_template_string(question_template, question_texts=question_texts, choice_texts=choice_texts)

@app.route("/submit", methods=["POST"])
def submit():
    student_id = session.get("student_id")
    name = session.get("name")
    email = session.get("email")

    total = 0
    detailed_results = []
    penalties = ""
    q10_penalty = 0

    correct_count = 0
    for key, correct in correct_answers.items():
        user_answer = request.form.get(key, "")
        is_correct = user_answer == correct
        if is_correct:
            total += 8
            correct_count += 1
        question_text = question_texts[key]
        user_choice = choice_texts[key].get(user_answer, "（未回答）")
        correct_choice = choice_texts[key][correct]
        detailed_results.append((question_text, user_choice, correct_choice, is_correct))

    answer10 = request.form.get("q10", "")
    if len(answer10) < 100:
        q10_penalty += 6
        penalties += "記述が100字以下: -6点\n"
    elif len(answer10) < 200:
        q10_penalty += 4
        penalties += "記述が200字以下: -4点\n"
    elif len(answer10) < 300:
        q10_penalty += 2
        penalties += "記述が300字以下: -2点\n"

    typo_pattern = r"[^\u3040-\u309F\u30A0-\u30FF\u4E00-\u9FAF\s。、．，]"
    typos = re.findall(typo_pattern, answer10)
    if typos:
        q10_penalty += len(typos)
        penalties += f"記号・誤字: -{len(typos)}点\n"

    q10_score = max(0, 28 - q10_penalty)
    final_score = total + q10_score

    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    if not os.path.exists(RESULTS_FILE):
        wb = Workbook()
        ws = wb.active
        ws.append(["学籍番号", "氏名", "メール", "選択式点", "正答数", "記述式点", "総合点", "受験時間", "記述減点", "問10回答"])
    else:
        wb = openpyxl.load_workbook(RESULTS_FILE)
        ws = wb.active

    ws.append([student_id, name, email, total, correct_count, q10_score, final_score, now, penalties, answer10])
    wb.save(RESULTS_FILE)

    result_html = "<h2>正誤一覧</h2><table border='1'><tr><th>問題</th><th>あなたの答え</th><th>正解</th><th>判定</th></tr>"
    for q, ua, ca, ok in detailed_results:
        result_html += f"<tr><td>{q}</td><td>{ua}</td><td>{ca}</td><td>{'〇' if ok else '×'}</td></tr>"
    result_html += "</table><br><a href='/'>戻る</a>"
    return result_html

@app.route("/admin", methods=["GET", "POST"])
def admin():
    if "admin_logged_in" not in session:
        if request.method == "POST":
            if request.form.get("password") == ADMIN_PASSWORD:
                session["admin_logged_in"] = True
                return redirect(url_for("admin"))
            else:
                return "<p style='color:red;'>パスワードが間違っています。</p>" + admin_login_form()
        return admin_login_form()

    if not os.path.exists(RESULTS_FILE):
        return "まだ受験者はいません。"

    wb = openpyxl.load_workbook(RESULTS_FILE)
    ws = wb.active
    rows = list(ws.iter_rows(min_row=2, values_only=True))

    html = """
    <h2>受験結果一覧</h2>
    <a href="/logout">ログアウト</a><br><br>
    <table border='1' cellpadding='5' style="border-collapse: collapse; font-size: 12px;">
      <tr>
        <th>学籍番号</th>
        <th style="min-width: 120px;">氏名</th>
        <th>メール</th>
        <th style="min-width: 60px;">選択式点</th>
        <th style="min-width: 50px;">正答数</th>
        <th style="min-width: 60px;">記述式点</th>
        <th style="min-width: 40px;">総合点</th>
        <th style="min-width: 80px;">受験時間</th>
        <th style="min-width: 100px;">記述減点内容</th>
        <th>問10回答</th>
      </tr>
    """

    typo_pattern = r"[^\u3040-\u309F\u30A0-\u30FF\u4E00-\u9FAF\s。、．，]"

    for row in rows:
        if len(row) < 10:
            continue
        student_id, name, email, choice_score, correct_count, q10_score, final_score, timestamp, penalty, q10_answer = row

        highlighted_q10 = ""
        for char in str(q10_answer):
            if re.match(typo_pattern, char):
                highlighted_q10 += f"<span style='color:red'>{char}</span>"
            else:
                highlighted_q10 += char

        html += f"""
        <tr>
          <td>{student_id}</td>
          <td>{name}</td>
          <td>{email}</td>
          <td>{choice_score}</td>
          <td>{correct_count} / 9</td>
          <td>{q10_score}</td>
          <td>{final_score}</td>
          <td>{timestamp}</td>
          <td><div style='white-space: pre-wrap; font-size: 11px;'>{penalty}</div></td>
          <td><div style='white-space: pre-wrap; font-size: 10px;'>{highlighted_q10}</div></td>
        </tr>
        """

    html += "</table><br><a href='/download_excel'>Excelをダウンロード</a>"
    return html

@app.route("/download_excel")
def download_excel():
    if not os.path.exists(RESULTS_FILE):
        return "ファイルが存在しません。"
    return send_file(RESULTS_FILE, as_attachment=True)

@app.route("/logout")
def logout():
    session.pop("admin_logged_in", None)
    return redirect(url_for("index"))

if __name__ == "__main__":
    app.run(debug=True, host="0.0.0.0")
